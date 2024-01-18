from django.shortcuts import render

# Create your views here.
from openpyxl import load_workbook
from .models import Place
import tablib

# TODO: use django_import_export
def import_from_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            id, code, capacity, rate, area_size = row
            Place.objects.create(id=id, code=code, capacity=capacity, rate=rate, area_size=area_size)

        return render(request, 'import_success.html')

    return render(request, 'import_form.html')